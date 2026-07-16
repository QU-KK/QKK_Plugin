# This program is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTIBILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program. If not, see <http://www.gnu.org/licenses/>.

bl_info = {
    "name" : "ZMD_Tool",
    "author" : "渠奎奎", 
    "description" : "",
    "blender" : (5, 2, 0),
    "version" : (1, 0, 0),
    "location" : "",
    "warning" : "",
    "doc_url": "", 
    "tracker_url": "", 
    "category" : "3D View" 
}


import bpy
import bpy.utils.previews
import openpyxl




def string_to_int(value):
    if value.isdigit():
        return int(value)
    return 0


def string_to_icon(value):
    if value in bpy.types.UILayout.bl_rna.functions["prop"].parameters["icon"].enum_items.keys():
        return bpy.types.UILayout.bl_rna.functions["prop"].parameters["icon"].enum_items[value].value
    return string_to_int(value)


addon_keymaps = {}
_icons = None
node_tree_001 = {'sna_zmd_mat_list': [], 'sna_zmd_img_list': [], }
class SNA_PT_ZMD_MAYA_019F5(bpy.types.Panel):
    bl_label = 'ZMD_互导Maya'
    bl_idname = 'SNA_PT_ZMD_MAYA_019F5'
    bl_space_type = 'VIEW_3D'
    bl_region_type = 'UI'
    bl_context = ''
    bl_category = 'ZMD'
    bl_order = 2
    bl_ui_units_x=0

    @classmethod
    def poll(cls, context):
        return not (False)

    def draw_header(self, context):
        layout = self.layout

    def draw(self, context):
        layout = self.layout
        col_90488 = layout.column(heading='', align=False)
        col_90488.alert = False
        col_90488.enabled = True
        col_90488.active = True
        col_90488.use_property_split = False
        col_90488.use_property_decorate = False
        col_90488.scale_x = 1.0
        col_90488.scale_y = 1.0
        col_90488.alignment = 'Expand'.upper()
        col_90488.operator_context = "INVOKE_DEFAULT" if True else "EXEC_DEFAULT"
        col_8D7CD = col_90488.column(heading='', align=False)
        col_8D7CD.alert = False
        col_8D7CD.enabled = True
        col_8D7CD.active = True
        col_8D7CD.use_property_split = False
        col_8D7CD.use_property_decorate = False
        col_8D7CD.scale_x = 1.0
        col_8D7CD.scale_y = 1.5
        col_8D7CD.alignment = 'Expand'.upper()
        col_8D7CD.operator_context = "INVOKE_DEFAULT" if True else "EXEC_DEFAULT"
        op = col_8D7CD.operator('sna.maya_to_blender_1f248', text='导入至Blender', icon_value=string_to_icon('IMPORT'), emboss=True, depress=False)
        op = col_8D7CD.operator('sna.blender_to_maya_4a904', text='导出至Maya', icon_value=string_to_icon('EXPORT'), emboss=True, depress=False)
        col_90488.prop(bpy.context.scene, 'sna_zmd_tex_export', text='贴图导出（Color、NRO）', icon_value=0, emboss=True)
        col_90488.prop(bpy.context.scene, 'sna_zmd_obj_mat_name', text='模型、材质  +1 → _1', icon_value=0, emboss=True)


class SNA_OT_Blender_To_Maya_4A904(bpy.types.Operator):
    bl_idname = "sna.blender_to_maya_4a904"
    bl_label = "Blender_To_Maya"
    bl_description = ""
    bl_options = {"REGISTER", "UNDO"}

    @classmethod
    def poll(cls, context):
        if bpy.app.version >= (3, 0, 0) and True:
            cls.poll_message_set('')
        return not False

    def execute(self, context):
        zmd_tex_export = bpy.context.scene.sna_zmd_tex_export
        zmd_obj_mat_name = bpy.context.scene.sna_zmd_obj_mat_name
        Path = 'C:\Blender_Cache\BlenderToMaya\Qkk_BlenderToMaya.FBX'

        def replace_names(old_str, new_str):
            if zmd_obj_mat_name:
                # 获取当前选中的所有物体
                selected_objects = bpy.context.selected_objects
                if not selected_objects:
                    print("请先在场景中选中至少一个物体！")
                    return
                # 用于记录已被修改过材质的集合，防止同一个材质被重复处理
                processed_materials = set()
                for obj in selected_objects:
                    # 1. 替换物体名称
                    if old_str in obj.name:
                        old_obj_name = obj.name
                        obj.name = obj.name.replace(old_str, new_str)
                        print(f"物体改名: {old_obj_name} -> {obj.name}")
                    # 2. 遍历物体关联的材质槽
                    if obj.material_slots:
                        for slot in obj.material_slots:
                            mat = slot.material
                            # 确保材质槽里确实有材质，且该材质还没被处理过
                            if mat and mat not in processed_materials:
                                processed_materials.add(mat)
                                if old_str in mat.name:
                                    old_mat_name = mat.name
                                    mat.name = mat.name.replace(old_str, new_str)
                                    print(f"  材质改名: {old_mat_name} -> {mat.name}")

        def mat_tex(switch):
            if zmd_tex_export:
                #处理材质贴图链接
                selected_meshes = [obj for obj in bpy.context.selected_objects if obj.type == 'MESH']
                unique_materials = {
                    slot.material 
                    for obj in selected_meshes 
                    for slot in obj.material_slots 
                    if slot.material is not None
                }
                # 转换为最终的列表
                materials_list = list(unique_materials)
                for mat in materials_list:
                    nodes = mat.node_tree.nodes
                    if nodes.get('Shader'):
                        # 获取材质
                        nodes = mat.node_tree.nodes
                        links = mat.node_tree.links
                        # 清理节点
                        nodes_to_delete = ["ZMD_贴图代理", "ZMD_法线贴图"]
                        for node in list(nodes):
                                if node.name in nodes_to_delete:
                                   nodes.remove(node)
                        node_out = nodes.get('ZMD_材质输出')
                        node_shader = nodes.get('Shader')  
                        if switch:
                            # 创建PrincipledBSDF
                            node_principled = nodes.new(type='ShaderNodeBsdfPrincipled')
                            node_principled.name = "ZMD_贴图代理"   # 内部系统名称
                            node_principled.label = "ZMD_贴图代理"  # 界面显示标签
                            # 创建NormalMap
                            node_normal = nodes.new(type='ShaderNodeNormalMap')
                            node_normal.name = "ZMD_法线贴图"   # 内部系统名称
                            node_normal.label = "ZMD_法线贴图"  # 界面显示标签
                            # 执行连线
                            links.new(node_normal.outputs['Normal'], node_principled.inputs['Normal'])
                            node_color = nodes.get('Color')
                            node_nro = nodes.get('NRO')
                            links.new(node_color.outputs['Color'], node_principled.inputs['Base Color'])
                            links.new(node_nro.outputs['Color'], node_normal.inputs['Color'])
                            links.new(node_principled.outputs['BSDF'], node_out.inputs['Surface'])
                        else:
                            links.new(node_shader.outputs['BSDF'], node_out.inputs['Surface'])
                    else:
                        print('没有相关材质不进行处理')
        #切换至fbx贴图链接模式
        mat_tex(True)
        replace_names('+1_','_1_')
        #导出fbx
        bpy.ops.export_scene.fbx(
        filepath=Path,#模型路径
        #包括
        use_selection=True,#选择项
        use_visible=False, #可见项
        use_active_collection=False,#激活的集合
        object_types={'MESH','EMPTY'},#数据类似
        use_custom_props=True,#自定义属性
        #变换
        global_scale=1.0,#缩放
        apply_scale_options='FBX_SCALE_NONE',#应用缩放
        axis_forward='Y',#向前
        axis_up='Z',#向上
        apply_unit_scale=True,#应用单位
        use_space_transform=True,#使用空间变换
        bake_space_transform=False,#应用变换
        #几何数据
        mesh_smooth_type='EDGE',#平滑
        use_subsurf=False,#导出表面细分
        use_mesh_modifiers=True,#应用修改器
        use_mesh_edges=False,#松散边
        use_triangles=False,#三角面
        use_tspace=False,#切向空间
        colors_type='SRGB',#顶点色空间
        prioritize_active_color=False,#活动颜色优先
        #动画
        bake_anim=False,#动画
        #其他
        path_mode='ABSOLUTE',#路径模式
        embed_textures=False,#内嵌纹理
        batch_mode='OFF',#批量模式
        )
        #切换至shader模式
        mat_tex(False)
        replace_names('_1_','+1_')
        self.report({'INFO'}, message='OK!')
        return {"FINISHED"}

    def invoke(self, context, event):
        return self.execute(context)


class SNA_OT_Maya_To_Blender_1F248(bpy.types.Operator):
    bl_idname = "sna.maya_to_blender_1f248"
    bl_label = "Maya_To_Blender"
    bl_description = ""
    bl_options = {"REGISTER", "UNDO"}

    @classmethod
    def poll(cls, context):
        if bpy.app.version >= (3, 0, 0) and True:
            cls.poll_message_set('')
        return not False

    def execute(self, context):
        Path = 'C:\Blender_Cache\BlenderToMaya\Qkk_MayaToBlender.FBX'
        bpy.ops.wm.fbx_import(
        #'INVOKE_DEFAULT',
        filepath=Path,
        # 常规
        global_scale=1.0,#缩放
        use_custom_props=False, #自定义属性
        use_custom_props_enum_as_string=False, #导入枚举为字符串
        # 几何数据
        use_custom_normals=True, #自定义法向
        import_subdivision=False, #细分曲面
        import_colors='SRGB', #顶点色
        validate_meshes=True, #检查网格
        # 材质引用
        mtl_name_collision_mode='REFERENCE_EXISTING',#引用同名材质
        # 动画
        use_anim=False, #动画
        )
        self.report({'INFO'}, message='OK!')
        return {"FINISHED"}

    def invoke(self, context, event):
        return self.execute(context)


class SNA_PT_ZMD__F55C4(bpy.types.Panel):
    bl_label = 'ZMD_批量处理'
    bl_idname = 'SNA_PT_ZMD__F55C4'
    bl_space_type = 'VIEW_3D'
    bl_region_type = 'UI'
    bl_context = ''
    bl_category = 'ZMD'
    bl_order = 5
    bl_ui_units_x=0

    @classmethod
    def poll(cls, context):
        return not (False)

    def draw_header(self, context):
        layout = self.layout

    def draw(self, context):
        layout = self.layout
        split_8F6E3 = layout.split(factor=0.5, align=True)
        split_8F6E3.alert = False
        split_8F6E3.enabled = True
        split_8F6E3.active = True
        split_8F6E3.use_property_split = False
        split_8F6E3.use_property_decorate = False
        split_8F6E3.scale_x = 1.0
        split_8F6E3.scale_y = 1.5
        split_8F6E3.alignment = 'Expand'.upper()
        if not True: split_8F6E3.operator_context = "EXEC_DEFAULT"
        op = split_8F6E3.operator('sna.zmd_name_b2a11', text='+1_  →  _1_', icon_value=0, emboss=True, depress=False)
        op.sna_old_str = '+1_'
        op.sna_new_str = '_1_'
        op = split_8F6E3.operator('sna.zmd_name_b2a11', text='_1_  →  +1_', icon_value=0, emboss=True, depress=False)
        op.sna_old_str = '_1_'
        op.sna_new_str = '+1_'
        split_98250 = layout.split(factor=0.5, align=True)
        split_98250.alert = False
        split_98250.enabled = True
        split_98250.active = True
        split_98250.use_property_split = False
        split_98250.use_property_decorate = False
        split_98250.scale_x = 1.0
        split_98250.scale_y = 1.5
        split_98250.alignment = 'Expand'.upper()
        if not True: split_98250.operator_context = "EXEC_DEFAULT"
        op = split_98250.operator('sna.my_generic_operator_d747e', text='打组', icon_value=0, emboss=True, depress=False)
        op = split_98250.operator('sna.my_generic_operator_c0922', text='拆组', icon_value=0, emboss=True, depress=False)
        col_B5A6D = layout.column(heading='', align=False)
        col_B5A6D.alert = False
        col_B5A6D.enabled = True
        col_B5A6D.active = True
        col_B5A6D.use_property_split = False
        col_B5A6D.use_property_decorate = False
        col_B5A6D.scale_x = 1.0
        col_B5A6D.scale_y = 1.5
        col_B5A6D.alignment = 'Expand'.upper()
        col_B5A6D.operator_context = "INVOKE_DEFAULT" if True else "EXEC_DEFAULT"
        op = col_B5A6D.operator('sna.my_generic_operator_d6702', text='全量材质名称规范化', icon_value=0, emboss=True, depress=False)


class SNA_OT_Zmd_Name_B2A11(bpy.types.Operator):
    bl_idname = "sna.zmd_name_b2a11"
    bl_label = "ZMD_Name"
    bl_description = ""
    bl_options = {"REGISTER", "UNDO"}
    sna_old_str: bpy.props.StringProperty(name='old_str', description='', options={'HIDDEN'}, default='', subtype='NONE', maxlen=0)
    sna_new_str: bpy.props.StringProperty(name='new_str', description='', options={'HIDDEN'}, default='', subtype='NONE', maxlen=0)

    @classmethod
    def poll(cls, context):
        if bpy.app.version >= (3, 0, 0) and True:
            cls.poll_message_set('')
        return not False

    def execute(self, context):
        old_str = self.sna_old_str
        new_str = self.sna_new_str

        def replace_names(old_str, new_str):
            # 获取当前选中的所有物体
            selected_objects = bpy.context.selected_objects
            if not selected_objects:
                print("请先在场景中选中至少一个物体！")
                return
            # 用于记录已被修改过材质的集合，防止同一个材质被重复处理
            processed_materials = set()
            for obj in selected_objects:
                # 1. 替换物体名称
                if old_str in obj.name:
                    old_obj_name = obj.name
                    obj.name = obj.name.replace(old_str, new_str)
                    print(f"物体改名: {old_obj_name} -> {obj.name}")
                # 2. 遍历物体关联的材质槽
                if obj.material_slots:
                    for slot in obj.material_slots:
                        mat = slot.material
                        # 确保材质槽里确实有材质，且该材质还没被处理过
                        if mat and mat not in processed_materials:
                            processed_materials.add(mat)
                            if old_str in mat.name:
                                old_mat_name = mat.name
                                mat.name = mat.name.replace(old_str, new_str)
                                print(f"  材质改名: {old_mat_name} -> {mat.name}")
        replace_names(old_str,new_str)
        self.report({'INFO'}, message='OK！')
        return {"FINISHED"}

    def invoke(self, context, event):
        return self.execute(context)


class SNA_OT_My_Generic_Operator_D747E(bpy.types.Operator):
    bl_idname = "sna.my_generic_operator_d747e"
    bl_label = "打组"
    bl_description = ""
    bl_options = {"REGISTER", "UNDO"}

    @classmethod
    def poll(cls, context):
        if bpy.app.version >= (3, 0, 0) and True:
            cls.poll_message_set('')
        return not False

    def execute(self, context):
        #冻结变换
        bpy.ops.object.transform_apply(location=False, rotation=True, scale=True)
        #_lod0储存
        obj_lod0 = []
        for obj in bpy.context.selected_objects:
            obj_name = obj.name
            if '_lod0' in obj_name:
                obj_lod0.append(obj)
        PRE_Empty_List = []
        #创建结构体
        for obj in obj_lod0:
            #激活物体集合位置
            Collection = obj.users_collection[0]
            #创建PRE主空物体
            PRE_Name = obj.name.replace("_lod0", "")
            PRE_Empty = bpy.data.objects.new(PRE_Name, None)
            Collection.objects.link(PRE_Empty)
            PRE_Empty.location = obj.location
            PRE_Empty_List.append(PRE_Empty)
            #创建结构空物体
            LOD_Empty = bpy.data.objects.new('LOD__' + PRE_Name, None)
            COL1_Empty = bpy.data.objects.new('COL1__' + PRE_Name, None)
            COL3_Empty = bpy.data.objects.new('COL3__' + PRE_Name, None)
            shadowProxy_Empty = bpy.data.objects.new('shadowProxy__' + PRE_Name, None)
            HLOD_Empty = bpy.data.objects.new('HLOD__' + PRE_Name, None)
            #空物体
            Empty_List = [LOD_Empty,COL1_Empty,COL3_Empty,shadowProxy_Empty,HLOD_Empty]
            #创建空物体并整理好结构
            for Empty in Empty_List:
                Collection.objects.link(Empty)
                Empty.location = (0,0,0)
                Empty.parent = PRE_Empty
            #模型放入结构
            for obj in bpy.context.selected_objects:
                if PRE_Name in obj.name:
                    if any(lod in obj.name for lod in ['_lod0', '_lod1', '_lod2', '_lod3', '_lod4']):
                        obj.parent = LOD_Empty
                        obj.location = (0,0,0)
                    if '_COL1_'in obj.name:
                        obj.parent = COL1_Empty
                        obj.location = (0,0,0)
                    if '_COL3_'in obj.name:
                        obj.parent = COL3_Empty
                        obj.location = (0,0,0)
                    if '_shadowProxy'in obj.name:
                        obj.parent = shadowProxy_Empty
                        obj.location = (0,0,0)
                    if '_HLOD'in obj.name:
                        obj.parent = HLOD_Empty
                        obj.location = (0,0,0)
        #选中最终结构
        bpy.ops.object.select_all(action='DESELECT')#取消全部选择
        for Empty in PRE_Empty_List:
            Empty.select_set(True)
            bpy.context.view_layer.objects.active = Empty
            bpy.ops.object.select_grouped(extend=True,type='CHILDREN_RECURSIVE')
        #删除子集为0的空物体
        for obj in bpy.context.selected_objects:    
            if obj.type == 'EMPTY' and len(obj.children) == 0:
                print(len(obj.children))
                bpy.data.objects.remove(obj)
        self.report({'INFO'}, message='OK！')
        return {"FINISHED"}

    def invoke(self, context, event):
        return self.execute(context)


class SNA_OT_My_Generic_Operator_C0922(bpy.types.Operator):
    bl_idname = "sna.my_generic_operator_c0922"
    bl_label = "拆组"
    bl_description = ""
    bl_options = {"REGISTER", "UNDO"}

    @classmethod
    def poll(cls, context):
        if bpy.app.version >= (3, 0, 0) and True:
            cls.poll_message_set('')
        return not False

    def execute(self, context):
        for obj in bpy.context.selected_objects:
            if obj.type == 'EMPTY':
                bpy.context.view_layer.objects.active = obj
                bpy.ops.object.select_grouped(extend=True,type='PARENT')
                bpy.ops.object.select_grouped(extend=True,type='CHILDREN_RECURSIVE')
        for obj in bpy.context.selected_objects:
            if obj.type == 'EMPTY':        
                bpy.ops.object.parent_clear(type='CLEAR_KEEP_TRANSFORM')        
                bpy.data.objects.remove(obj)
        self.report({'INFO'}, message='OK！')
        return {"FINISHED"}

    def invoke(self, context, event):
        return self.execute(context)


class SNA_OT_My_Generic_Operator_D6702(bpy.types.Operator):
    bl_idname = "sna.my_generic_operator_d6702"
    bl_label = "处理材质名称"
    bl_description = ""
    bl_options = {"REGISTER", "UNDO"}

    @classmethod
    def poll(cls, context):
        if bpy.app.version >= (3, 0, 0) and True:
            cls.poll_message_set('')
        return not False

    def execute(self, context):
        # 将所有名称中包含 '_1_' 的材质提取到一个列表中，避免在遍历时修改名称导致异常
        materials_to_process = [mat for mat in bpy.data.materials if '_1_' in mat.name]
        # 记录操作次数
        renamed_count = 0
        replaced_count = 0
        for mat in materials_to_process:
            old_name = mat.name
            # 生成目标名称
            target_name = old_name.replace('_1_', '+1_')
            # 检查目标名称是否已经存在于材质库中
            if target_name not in bpy.data.materials:
                # 情况 1：不重名，直接改名
                mat.name = target_name
                print(f"已重命名: {old_name} -> {target_name}")
                renamed_count += 1
            else:
                # 情况 2：发生重名，获取已经存在的目标材质
                target_mat = bpy.data.materials[target_name]
                print(f"发现重名: [{target_name}] 已存在。正在将使用 [{old_name}] 的模型替换为目标材质...")
                # 遍历场景中的所有物体
                for obj in bpy.data.objects:
                    # 只有支持材质的物体（网格、曲线、字体等）才有 material_slots 属性
                    if hasattr(obj, 'material_slots'):
                        for slot in obj.material_slots:
                            # 如果该物体的材质槽正在使用我们要被替换的材质
                            if slot.material == mat:
                                slot.material = target_mat
                                print(f"  - 物体 [{obj.name}] 的材质已被替换")
                replaced_count += 1
        print("-" * 30)
        print(f"操作完成！")
        print(f"直接重命名的材质数量: {renamed_count}")
        print(f"因重名而重新分配的材质数量: {replaced_count}")
        self.report({'INFO'}, message='OK!')
        return {"FINISHED"}

    def invoke(self, context, event):
        return self.execute(context)


class SNA_PT_excel_mat_D66CA(bpy.types.Panel):
    bl_label = '材质读表'
    bl_idname = 'SNA_PT_excel_mat_D66CA'
    bl_space_type = 'VIEW_3D'
    bl_region_type = 'UI'
    bl_context = ''
    bl_category = 'ZMD'
    bl_order = 4
    bl_ui_units_x=0

    @classmethod
    def poll(cls, context):
        return not (False)

    def draw_header(self, context):
        layout = self.layout

    def draw(self, context):
        layout = self.layout
        col_8B843 = layout.column(heading='', align=False)
        col_8B843.alert = False
        col_8B843.enabled = True
        col_8B843.active = True
        col_8B843.use_property_split = False
        col_8B843.use_property_decorate = False
        col_8B843.scale_x = 1.0
        col_8B843.scale_y = 1.5
        col_8B843.alignment = 'Expand'.upper()
        col_8B843.operator_context = "INVOKE_DEFAULT" if True else "EXEC_DEFAULT"
        op = col_8B843.operator('sna.read_excle_73eb5', text='读取材质', icon_value=0, emboss=True, depress=False)
        layout.prop(bpy.context.scene, 'sna_zmd_mat_excel_path', text='', icon_value=0, emboss=True)


class SNA_OT_Read_Excle_73Eb5(bpy.types.Operator):
    bl_idname = "sna.read_excle_73eb5"
    bl_label = "Read_Excle"
    bl_description = ""
    bl_options = {"REGISTER", "UNDO"}

    @classmethod
    def poll(cls, context):
        if bpy.app.version >= (3, 0, 0) and True:
            cls.poll_message_set('')
        return not False

    def execute(self, context):
        file_path = bpy.path.abspath(bpy.context.scene.sna_zmd_mat_excel_path)
        import os
        # 清理缺失图像

        def Remove_Missing_Img():
            # 统计删除了多少张图片
            deleted_count = 0
            # 遍历Blender文件中的所有图片数据
            for img in list(bpy.data.images):
                # 将 Blender 的相对路径（如 //text.png）转换为绝对路径
                abs_path = bpy.path.abspath(img.filepath)
                # 检查文件在硬盘上是否存在
                if not os.path.exists(abs_path):
                    print(f"检测到丢失的图片: {img.name} -> 路径: {abs_path}")
                    # 从Blender的数据块中彻底移除该图片
                    bpy.data.images.remove(img)
                    deleted_count += 1
            print(f"清理完成！共删除了 {deleted_count} 个丢失的图片数据块。")
        # 配置Shader材质

        def Set_Shader(Mat,Shader_Name):
            #Shader_Name = 'ZMD_Lit_Two'
            Blender_Path = 'C:\\Blender_Shader\\ZMD_Lit_Two.blend\\Material\\'
            # 清理重名的旧材质，防止冲突
            Old_Mat = bpy.data.materials.get(Shader_Name)
            if Old_Mat:
                bpy.data.materials.remove(Old_Mat)
            # 追加外部材质并修正节点组引用
            bpy.ops.wm.append(directory=Blender_Path, filename=Shader_Name, link=False)
            New_Mat = bpy.data.materials.get(Shader_Name)
            Node_Name = New_Mat.node_tree.nodes['Shader'].node_tree.name
            if Node_Name != Shader_Name:
                New_Mat.node_tree.nodes['Shader'].node_tree = bpy.data.node_groups[Shader_Name]
                bpy.data.node_groups.remove(bpy.data.node_groups[Node_Name])
            # 创建载体
            bpy.ops.mesh.primitive_cone_add()
            obj = bpy.context.active_object
            mesh = obj.data
            bpy.ops.object.material_slot_add()
            obj.material_slots[0].material = New_Mat
            # 切换至材质编辑器，复制新材质的节点
            bpy.context.area.ui_type = 'ShaderNodeTree'
            bpy.ops.node.select_all(action='SELECT')
            bpy.ops.node.clipboard_copy()
            # 删除临时材质
            bpy.context.blend_data.materials.remove(material=New_Mat)
            # 清空原材质节点,将复制的节点粘贴到原材质中，并切回3D视图
            Mat.node_tree.nodes.clear()
            obj.material_slots[0].material = Mat
            bpy.context.area.ui_type = 'ShaderNodeTree'
            bpy.ops.node.select_all(action='SELECT')
            bpy.ops.node.clipboard_paste()
            bpy.context.area.ui_type = 'VIEW_3D'
            # 删除载体
            bpy.context.blend_data.objects.remove(object=obj)
            bpy.context.blend_data.meshes.remove(mesh=mesh)

        def Img_Space():
            # 直接遍历 Blender 里的所有图像文件
            for img in bpy.data.images:
                # 统一将A通道设置为 “通道打包”
                img.alpha_mode = 'CHANNEL_PACKED'
                # 直接使用原始名称进行字符串包含判断
                if ("_N." in img.name) or ("_NRO." in img.name) or ("_M." in img.name):
                    img.colorspace_settings.name = 'Non-Color'
                    print(f"贴图 [{img.name}] 已直接设置为: Non-Color")
                elif ("_D." in img.name) or ("_E." in img.name):
                    img.colorspace_settings.name = 'sRGB'
                    print(f"贴图 [{img.name}] 已直接设置为: sRGB")
        print('Start')
        Remove_Missing_Img()
        # 加载Excel文件
        #file_path = 
        workbook = openpyxl.load_workbook(file_path,data_only=True) # data_only=True 获取单元格的值而非公式
        sheet = workbook.worksheets[0]
        # 添加 min_row=2 参数，直接从第二行开始读取，跳过第一行
        for row in sheet.iter_rows(min_row=2, max_col=10,values_only=True):
            # 如果第一列（索引为 0）为空，则跳过这一行
            if row[0] is not None:
                Mat = bpy.context.blend_data.materials.get(row[0])
                if Mat:
                    Shader_Name = row[1]
                    Set_Shader(Mat,Shader_Name)
                    # 贴图名称列表
                    Img_Name_List = [row[2],row[3],row[4],row[5],row[6]]
                    for Img_Name in Img_Name_List:
                        # 导入贴图
                        img = bpy.data.images.get(str(Img_Name) + '.tga')
                        image_path = os.path.dirname(file_path ) + '\\' + str(Img_Name) + '.tga'
                        if os.path.exists(image_path) and img == None:
                            bpy.data.images.load(image_path)
                    Mat.node_tree.nodes['Color'].image = bpy.data.images.get(str(row[2]) + '.tga')
                    Mat.node_tree.nodes['NRO'].image = bpy.data.images.get(str(row[3]) + '.tga')
                    Mat.node_tree.nodes['本体法线'].image = bpy.data.images.get(str(row[4]) + '.tga')
                    Mat.node_tree.nodes['三通道_MASK'].image = bpy.data.images.get(str(row[5]) + '.tga')
                    Mat.node_tree.nodes['自发光'].image = bpy.data.images.get(str(row[6]) + '.tga')
        Img_Space()
        print('END')
        return {"FINISHED"}

    def draw(self, context):
        layout = self.layout
        col_C95DC = layout.column(heading='', align=False)
        col_C95DC.alert = False
        col_C95DC.enabled = True
        col_C95DC.active = True
        col_C95DC.use_property_split = False
        col_C95DC.use_property_decorate = False
        col_C95DC.scale_x = 1.0
        col_C95DC.scale_y = 1.0
        col_C95DC.alignment = 'Expand'.upper()
        col_C95DC.operator_context = "INVOKE_DEFAULT" if True else "EXEC_DEFAULT"
        row_D926A = col_C95DC.row(heading='', align=True)
        row_D926A.alert = False
        row_D926A.enabled = True
        row_D926A.active = True
        row_D926A.use_property_split = False
        row_D926A.use_property_decorate = False
        row_D926A.scale_x = 1.0
        row_D926A.scale_y = 1.5
        row_D926A.alignment = 'Expand'.upper()
        row_D926A.operator_context = "INVOKE_DEFAULT" if True else "EXEC_DEFAULT"
        row_D926A.prop(bpy.context.scene, 'sna_zmd_mat_img_check', text=bpy.context.scene.sna_zmd_mat_img_check, icon_value=0, emboss=True, expand=True)
        if True:
            col_9EF47 = col_C95DC.column(heading='', align=True)
            col_9EF47.alert = False
            col_9EF47.enabled = True
            col_9EF47.active = True
            col_9EF47.use_property_split = False
            col_9EF47.use_property_decorate = False
            col_9EF47.scale_x = 1.0
            col_9EF47.scale_y = 1.0
            col_9EF47.alignment = 'Expand'.upper()
            col_9EF47.operator_context = "INVOKE_DEFAULT" if True else "EXEC_DEFAULT"
            for i_F6838 in range(len((node_tree_001['sna_zmd_img_list'] if (bpy.context.scene.sna_zmd_mat_img_check == '图像') else node_tree_001['sna_zmd_mat_list']))):
                split_3FF5A = col_9EF47.split(factor=0.800000011920929, align=True)
                split_3FF5A.alert = False
                split_3FF5A.enabled = True
                split_3FF5A.active = True
                split_3FF5A.use_property_split = False
                split_3FF5A.use_property_decorate = False
                split_3FF5A.scale_x = 1.0
                split_3FF5A.scale_y = 1.0
                split_3FF5A.alignment = 'Expand'.upper()
                if not True: split_3FF5A.operator_context = "EXEC_DEFAULT"
                split_B5956 = split_3FF5A.split(factor=0.20000000298023224, align=True)
                split_B5956.alert = False
                split_B5956.enabled = True
                split_B5956.active = True
                split_B5956.use_property_split = False
                split_B5956.use_property_decorate = False
                split_B5956.scale_x = 1.0
                split_B5956.scale_y = 1.0
                split_B5956.alignment = 'Expand'.upper()
                if not True: split_B5956.operator_context = "EXEC_DEFAULT"
                box_45212 = split_B5956.box()
                box_45212.alert = ((node_tree_001['sna_zmd_img_list'] if (bpy.context.scene.sna_zmd_mat_img_check == '图像') else node_tree_001['sna_zmd_mat_list'])[i_F6838][0].split('-')[1] == 'True')
                box_45212.enabled = True
                box_45212.active = True
                box_45212.use_property_split = False
                box_45212.use_property_decorate = False
                box_45212.alignment = 'Expand'.upper()
                box_45212.scale_x = 1.0
                box_45212.scale_y = 1.0
                if not True: box_45212.operator_context = "EXEC_DEFAULT"
                box_45212.label(text=(node_tree_001['sna_zmd_img_list'] if (bpy.context.scene.sna_zmd_mat_img_check == '图像') else node_tree_001['sna_zmd_mat_list'])[i_F6838][0].split('-')[0], icon_value=0)
                box_CC3C6 = split_B5956.box()
                box_CC3C6.alert = False
                box_CC3C6.enabled = True
                box_CC3C6.active = True
                box_CC3C6.use_property_split = False
                box_CC3C6.use_property_decorate = False
                box_CC3C6.alignment = 'Expand'.upper()
                box_CC3C6.scale_x = 1.0
                box_CC3C6.scale_y = 1.0
                if not True: box_CC3C6.operator_context = "EXEC_DEFAULT"
                box_CC3C6.label(text=(node_tree_001['sna_zmd_img_list'] if (bpy.context.scene.sna_zmd_mat_img_check == '图像') else node_tree_001['sna_zmd_mat_list'])[i_F6838][1], icon_value=0)
                box_CD881 = split_3FF5A.box()
                box_CD881.alert = ((node_tree_001['sna_zmd_img_list'] if (bpy.context.scene.sna_zmd_mat_img_check == '图像') else node_tree_001['sna_zmd_mat_list'])[i_F6838][2].split('-')[1] == 'True')
                box_CD881.enabled = True
                box_CD881.active = True
                box_CD881.use_property_split = False
                box_CD881.use_property_decorate = False
                box_CD881.alignment = 'Expand'.upper()
                box_CD881.scale_x = 1.0
                box_CD881.scale_y = 1.0
                if not True: box_CD881.operator_context = "EXEC_DEFAULT"
                box_CD881.label(text=(node_tree_001['sna_zmd_img_list'] if (bpy.context.scene.sna_zmd_mat_img_check == '图像') else node_tree_001['sna_zmd_mat_list'])[i_F6838][2].split('-')[0], icon_value=0)

    def invoke(self, context, event):
        # 将所有名称中包含 '_1_' 的材质提取到一个列表中，避免在遍历时修改名称导致异常
        materials_to_process = [mat for mat in bpy.data.materials if '_1_' in mat.name]
        # 记录操作次数
        renamed_count = 0
        replaced_count = 0
        for mat in materials_to_process:
            old_name = mat.name
            # 生成目标名称
            target_name = old_name.replace('_1_', '+1_')
            # 检查目标名称是否已经存在于材质库中
            if target_name not in bpy.data.materials:
                # 情况 1：不重名，直接改名
                mat.name = target_name
                print(f"已重命名: {old_name} -> {target_name}")
                renamed_count += 1
            else:
                # 情况 2：发生重名，获取已经存在的目标材质
                target_mat = bpy.data.materials[target_name]
                print(f"发现重名: [{target_name}] 已存在。正在将使用 [{old_name}] 的模型替换为目标材质...")
                # 遍历场景中的所有物体
                for obj in bpy.data.objects:
                    # 只有支持材质的物体（网格、曲线、字体等）才有 material_slots 属性
                    if hasattr(obj, 'material_slots'):
                        for slot in obj.material_slots:
                            # 如果该物体的材质槽正在使用我们要被替换的材质
                            if slot.material == mat:
                                slot.material = target_mat
                                print(f"  - 物体 [{obj.name}] 的材质已被替换")
                replaced_count += 1
        print("-" * 30)
        print(f"操作完成！")
        print(f"直接重命名的材质数量: {renamed_count}")
        print(f"因重名而重新分配的材质数量: {replaced_count}")
        file_path = bpy.path.abspath(bpy.context.scene.sna_zmd_mat_excel_path)
        ZMD_Mat_List = None
        ZMD_Img_List = None
        import os
        ZMD_Mat_List = []
        ZMD_Img_List = []
        # 加载Excel文件
        #file_path = 
        workbook = openpyxl.load_workbook(file_path,data_only=True) # data_only=True 获取单元格的值而非公式
        sheet = workbook.worksheets[0]
        # 添加 min_row=2 参数，直接从第二行开始读取，跳过第一行
        for row in sheet.iter_rows(min_row=2, max_col=10,values_only=True):
            # 如果第一列（索引为 0）为空，则跳过这一行
            if row[0] is not None:
                Mat_Name = row[0]
                Mat_Exist = ''
                if bpy.context.blend_data.materials.get(row[0]):
                    Mat_Exist = '材质存在-False'
                else:
                    Mat_Exist = '材质缺失-True'
                Mat_Name_Check = ''
                if 'M_' in Mat_Name and '+1_' in Mat_Name:
                    Mat_Name_Check = '名称正常-False'
                else:
                    Mat_Name_Check = '名称异常-True'
                ZMD_Mat_List.append([Mat_Exist,Mat_Name,Mat_Name_Check])
                # 贴图名称列表
                Img_Name_List = [row[2],row[3],row[4],row[5],row[6]]
                for Img_Name in Img_Name_List:
                    if Img_Name is not None:
                        image_path = os.path.dirname(file_path ) + '\\' + str(Img_Name) + '.tga'            
                        Img_Exist = os.path.exists(image_path) 
                        Img_Exist = ''
                        if os.path.exists(image_path):
                            Img_Exist = '路径存在-False'
                        else:
                            Img_Exist = '路径缺失-True'
                        Img_Name_Check = ''
                        if 'T_' in Img_Name and '+1_' in Img_Name:
                            Img_Name_Check = '名称正常-False'
                        else:
                            Img_Name_Check = '名称异常-True'
                        ZMD_Img_List.append([Img_Exist,Img_Name,Img_Name_Check])
        node_tree_001['sna_zmd_mat_list'] = ZMD_Mat_List
        node_tree_001['sna_zmd_img_list'] = ZMD_Img_List
        return context.window_manager.invoke_props_dialog(self, width=450)


def register():
    global _icons
    _icons = bpy.utils.previews.new()
    bpy.types.Scene.sna_zmd_tex_export = bpy.props.BoolProperty(name='zmd_tex_export', description='', default=False)
    bpy.types.Scene.sna_zmd_obj_mat_name = bpy.props.BoolProperty(name='zmd_obj_mat_name', description='', default=True)
    bpy.types.Scene.sna_zmd_mat_excel_path = bpy.props.StringProperty(name='zmd_mat_excel_path', description='', default='', subtype='FILE_PATH', maxlen=0)
    bpy.types.Scene.sna_zmd_mat_img_check = bpy.props.EnumProperty(name='zmd_mat_img_check', description='', items=[('材质', '材质', '', 0, 0), ('图像', '图像', '', 0, 1)])
    bpy.utils.register_class(SNA_PT_ZMD_MAYA_019F5)
    bpy.utils.register_class(SNA_OT_Blender_To_Maya_4A904)
    bpy.utils.register_class(SNA_OT_Maya_To_Blender_1F248)
    bpy.utils.register_class(SNA_PT_ZMD__F55C4)
    bpy.utils.register_class(SNA_OT_Zmd_Name_B2A11)
    bpy.utils.register_class(SNA_OT_My_Generic_Operator_D747E)
    bpy.utils.register_class(SNA_OT_My_Generic_Operator_C0922)
    bpy.utils.register_class(SNA_OT_My_Generic_Operator_D6702)
    bpy.utils.register_class(SNA_PT_excel_mat_D66CA)
    bpy.utils.register_class(SNA_OT_Read_Excle_73Eb5)


def unregister():
    global _icons
    bpy.utils.previews.remove(_icons)
    wm = bpy.context.window_manager
    kc = wm.keyconfigs.addon
    for km, kmi in addon_keymaps.values():
        km.keymap_items.remove(kmi)
    addon_keymaps.clear()
    del bpy.types.Scene.sna_zmd_mat_img_check
    del bpy.types.Scene.sna_zmd_mat_excel_path
    del bpy.types.Scene.sna_zmd_obj_mat_name
    del bpy.types.Scene.sna_zmd_tex_export
    bpy.utils.unregister_class(SNA_PT_ZMD_MAYA_019F5)
    bpy.utils.unregister_class(SNA_OT_Blender_To_Maya_4A904)
    bpy.utils.unregister_class(SNA_OT_Maya_To_Blender_1F248)
    bpy.utils.unregister_class(SNA_PT_ZMD__F55C4)
    bpy.utils.unregister_class(SNA_OT_Zmd_Name_B2A11)
    bpy.utils.unregister_class(SNA_OT_My_Generic_Operator_D747E)
    bpy.utils.unregister_class(SNA_OT_My_Generic_Operator_C0922)
    bpy.utils.unregister_class(SNA_OT_My_Generic_Operator_D6702)
    bpy.utils.unregister_class(SNA_PT_excel_mat_D66CA)
    bpy.utils.unregister_class(SNA_OT_Read_Excle_73Eb5)
